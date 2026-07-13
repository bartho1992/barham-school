from flask import render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_login import login_required
from models import db, Ecole, Eleve, Classe, Note, Paiement, AbonnementService, Document, Bulletin
from app import app, get_current_ecole_id
import os, io, openpyxl, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _annee_courante(e):
    return session.get('annee_scolaire', e.annee_scolaire if e else '')

@app.route('/eleves')
@login_required
def eleves():
    ecole_id = get_current_ecole_id()
    e = Ecole.query.get(ecole_id); classes = Classe.query.filter_by(ecole_id=ecole_id).all(); annee = _annee_courante(e)
    q = Eleve.query.filter_by(annee_scolaire=annee, ecole_id=ecole_id)
    fc = request.args.get('classe',''); fs = request.args.get('search','')
    if fc: q = q.filter_by(classe_id=fc)
    if fs: q = q.filter((Eleve.prenom.contains(fs))|(Eleve.nom.contains(fs))|(Eleve.code.contains(fs)))
    return render_template('eleves/index.html', eleves=q.order_by(Eleve.nom,Eleve.prenom).all(), classes=classes, ecole=e, filter_classe=fc, filter_search=fs)

@app.route('/api/eleves/search')
@login_required
def api_eleves_search():
    ecole_id = get_current_ecole_id()
    q = request.args.get('q', '')
    if not q: return jsonify([])
    els = Eleve.query.filter_by(ecole_id=ecole_id).filter((Eleve.prenom.contains(q))|(Eleve.nom.contains(q))|(Eleve.code.contains(q))).limit(10).all()
    return jsonify([{'id': e.id, 'prenom': e.prenom, 'nom': e.nom, 'code': e.code, 'classe': e.classe.nom if e.classe else ''} for e in els])

@app.route('/eleves/ajouter', methods=['GET','POST'])
@login_required
def eleve_ajouter():
    ecole_id = get_current_ecole_id()
    e = Ecole.query.get(ecole_id); classes = Classe.query.filter_by(ecole_id=ecole_id).all(); annee = _annee_courante(e)
    
    # Generate automatic code
    last_el = Eleve.query.filter_by(ecole_id=ecole_id).order_by(Eleve.id.desc()).first()
    next_id = (last_el.id + 1) if last_el else 1
    from datetime import datetime
    year = datetime.now().year
    auto_code = f"{year}{next_id:04d}"
    
    if request.method == 'POST':
        code = request.form.get('code') or auto_code
        if Eleve.query.filter_by(code=code, ecole_id=ecole_id).first() and not request.form.get('id'):
            flash('Code existe déjà pour cet établissement','danger')
            return render_template('eleves/form.html', classes=classes, ecole=e, auto_code=auto_code)
        
        el = Eleve(code=code, prenom=request.form.get('prenom'), nom=request.form.get('nom'),
            sexe=request.form.get('sexe'), classe_id=request.form.get('classe_id') or None,
            tel=request.form.get('tel'), date_naissance=request.form.get('date_naissance'),
            lieu_naissance=request.form.get('lieu_naissance'), tuteur=request.form.get('tuteur'),
            adresse=request.form.get('adresse'), precedente_ecole=request.form.get('precedente_ecole'),
            date_entree=request.form.get('date_entree'), observations=request.form.get('observations'),
            situation=request.form.get('situation','Inscrit'), annee_scolaire=annee, ecole_id=ecole_id)
        db.session.add(el); db.session.commit()
        # Gérer la photo
        photo_file = request.files.get('photo')
        if photo_file and photo_file.filename:
            ext = photo_file.filename.rsplit('.', 1)[-1].lower() if '.' in photo_file.filename else 'jpg'
            if ext in ('jpg','jpeg','png','gif','bmp','webp'):
                filename = f"eleve_{el.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
                upload_dir = os.path.join(os.path.dirname(__file__), 'static', 'photos')
                os.makedirs(upload_dir, exist_ok=True)
                photo_file.save(os.path.join(upload_dir, filename))
                el.photo = filename
                db.session.commit()
        # Mettre à jour l'effectif de la classe
        if el.classe_id:
            cls = Classe.query.get(el.classe_id)
            if cls: cls.effectif = Eleve.query.filter_by(classe_id=el.classe_id).count()
        db.session.commit()
        flash('Élève ajouté','success'); return redirect(url_for('eleves'))
    return render_template('eleves/form.html', classes=classes, ecole=e, auto_code=auto_code)

@app.route('/eleves/modifier/<int:id>', methods=['GET','POST'])
@login_required
def eleve_modifier(id):
    ecole_id = get_current_ecole_id()
    e = Ecole.query.get(ecole_id); el = Eleve.query.get_or_404(id); classes = Classe.query.filter_by(ecole_id=ecole_id).all()
    if el.ecole_id != ecole_id: flash('Accès non autorisé','danger'); return redirect(url_for('eleves'))
    
    if request.method == 'POST':
        from datetime import datetime
        for f in ['code','prenom','nom','sexe','tel','date_naissance','lieu_naissance','tuteur','adresse','precedente_ecole','date_entree','observations','situation']:
            setattr(el, f, request.form.get(f))
        old_classe_id = el.classe_id
        el.classe_id = request.form.get('classe_id') or None
        db.session.commit()
        # Gérer la photo
        photo_file = request.files.get('photo')
        if photo_file and photo_file.filename:
            ext = photo_file.filename.rsplit('.', 1)[-1].lower() if '.' in photo_file.filename else 'jpg'
            if ext in ('jpg','jpeg','png','gif','bmp','webp'):
                filename = f"eleve_{el.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
                upload_dir = os.path.join(os.path.dirname(__file__), 'static', 'photos')
                os.makedirs(upload_dir, exist_ok=True)
                photo_file.save(os.path.join(upload_dir, filename))
                el.photo = filename
                db.session.commit()
        # Mettre à jour l'effectif des classes concernées
        for cid in set(filter(None, [old_classe_id, el.classe_id])):
            cls = Classe.query.get(cid)
            if cls: cls.effectif = Eleve.query.filter_by(classe_id=cid).count()
        db.session.commit()
        flash('Élève modifié','success'); return redirect(url_for('eleves'))
    return render_template('eleves/form.html', eleve=el, classes=classes, ecole=e)

@app.route('/eleves/supprimer/<int:id>', methods=['POST'])
@login_required
def eleve_supprimer(id):
    ecole_id = get_current_ecole_id()
    el = Eleve.query.get_or_404(id)
    if el.ecole_id != ecole_id: flash('Accès non autorisé','danger'); return redirect(url_for('eleves'))
    
    classe_id = el.classe_id
    # Supprimer les données liées
    Note.query.filter_by(eleve_id=id).delete()
    Paiement.query.filter_by(eleve_id=id).delete()
    AbonnementService.query.filter_by(eleve_id=id).delete()
    Document.query.filter_by(eleve_id=id).delete()
    Bulletin.query.filter_by(eleve_id=id).delete()
    db.session.delete(el)
    db.session.commit()
    # Mettre à jour l'effectif de la classe
    if classe_id:
        cls = Classe.query.get(classe_id)
        if cls: cls.effectif = Eleve.query.filter_by(classe_id=classe_id).count()
        db.session.commit()
    flash('Supprimé','success'); return redirect(url_for('eleves'))

@app.route('/eleves/supprimer-bulk', methods=['POST'])
@login_required
def eleves_supprimer_bulk():
    ecole_id = get_current_ecole_id()
    ids = request.form.getlist('eleve_ids[]')
    if not ids:
        flash('Aucun élève sélectionné', 'warning')
        return redirect(url_for('eleves'))
    count = 0
    for eid in [int(i) for i in ids]:
        el = Eleve.query.filter_by(id=eid, ecole_id=ecole_id).first()
        if not el:
            continue
        classe_id = el.classe_id
        Note.query.filter_by(eleve_id=eid).delete()
        Paiement.query.filter_by(eleve_id=eid).delete()
        AbonnementService.query.filter_by(eleve_id=eid).delete()
        Document.query.filter_by(eleve_id=eid).delete()
        Bulletin.query.filter_by(eleve_id=eid).delete()
        db.session.delete(el)
        if classe_id:
            cls = Classe.query.get(classe_id)
            if cls: cls.effectif = Eleve.query.filter_by(classe_id=classe_id).count()
        count += 1
    db.session.commit()
    flash(f'{count} élève(s) supprimé(s) avec succès', 'success')
    return redirect(url_for('eleves'))

@app.route('/eleves/fiche/<int:id>')
@login_required
def eleve_fiche(id):
    ecole_id = get_current_ecole_id()
    e = Ecole.query.get(ecole_id); el = Eleve.query.get_or_404(id); annee = _annee_courante(e)
    if el.ecole_id != ecole_id: flash('Accès non autorisé','danger'); return redirect(url_for('eleves'))
    
    return render_template('eleves/fiche.html', eleve=el, ecole=e, 
        notes=Note.query.filter_by(eleve_id=id, annee_scolaire=annee).all(), 
        paiements=Paiement.query.filter_by(eleve_id=id, annee_scolaire=annee).all())


@app.route('/eleves/importer', methods=['POST'])
@login_required
def eleves_importer():
    """Importe des eleves depuis un fichier Excel"""
    ecole_id = get_current_ecole_id()
    e = Ecole.query.get(ecole_id); annee = _annee_courante(e)
    classes = {c.nom.lower().strip(): c for c in Classe.query.filter_by(ecole_id=ecole_id).all()}
    
    f = request.files.get('fichier_excel')
    if not f or f.filename == '':
        flash('Aucun fichier selectionne', 'danger')
        return redirect(url_for('eleves'))
    
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as ex:
        flash(f'Erreur lecture fichier: {str(ex)}', 'danger')
        return redirect(url_for('eleves'))
    
    def _normaliser(texte):
        """Normalise un texte: minuscules, sans accents, sans stop words, sans apostrophes"""
        t = unicodedata.normalize('NFKD', texte.strip().lower()).encode('ASCII', 'ignore').decode()
        for mot in [' de ', ' d\'', ' du ', ' de l\'', ' la ', ' le ', ' a ', ' au ', '-']:
            t = t.replace(mot, ' ')
        return t.replace('\'', '').replace('(', '').replace(')', '').replace('/', '_').replace(' ', '_').strip('_').replace('__', '_')

    # --- Detection intelligente des colonnes (row 1, 2, 3) ---
    headers = {}       # cle normalisee avec underscores: {'prenom': 1, 'nom': 2, ...}
    headers_bruts = {}  # cle normalisee avec stop words: {'prenom': 1, ...}
    col_names_raw = []  # noms bruts pour diagnostic
    
    for row_check in range(1, 4):  # Chercher dans les 3 premieres lignes
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row_check, col).value
            if val:
                cle_brute = unicodedata.normalize('NFKD', str(val).strip().lower()).encode('ASCII', 'ignore').decode().replace('\'', '').replace(' ', '_').replace('(', '').replace(')', '')
                cle_norm = _normaliser(str(val))
                if cle_brute not in headers:
                    headers[cle_brute] = col
                    headers_bruts[cle_norm] = col
                    col_names_raw.append(str(val).strip())
    
    # --- Detection automatique avec listes de synonymes ---
    # Chaque champ a une liste de mots-cles possibles, classes par priorite
    SYNONYMES = {
        'prenom': ['prenom', 'prenoms', 'first_name', 'firstname', 'first name', 'given_name', 'givenname'],
        'nom': ['nom', 'name', 'last_name', 'lastname', 'last name', 'surname', 'family_name', 'familyname', 'nom_de_famille', 'nom_famille'],
        'sexe': ['sexe', 'genre', 'gender', 'sex', 'civilite', 'civilite', 'm_f', 'mf'],
        'classe': ['classe', 'class', 'niveau', 'grade', 'level', 'form', 'division', 'div'],
        'tel': ['tel', 'telephone', 'phone', 'mobile', 'cell', 'contact', 'gsm', 'phone_number', 'phonenumber', 'numero'],
        'date_naissance': ['date_naissance', 'date_de_naissance', 'date naissance', 'ne_le', 'ne le', 'ne(e)_le', 'birth_date', 'birthdate', 'birth', 'anniversaire', 'date_of_birth', 'dob', 'naissance', 'date_naiss'],
        'lieu_naissance': ['lieu_naissance', 'lieu_de_naissance', 'lieu naissance', 'ne_a', 'ne a', 'birth_place', 'birthplace', 'place_of_birth', 'pob', 'ville_naissance'],
        'tuteur': ['tuteur', 'parent', 'responsable', 'responsable_legal', 'gardien', 'guardian', 'pere', 'mere', 'pere_mere', 'father', 'mother', 'tuteur_legal'],
        'adresse': ['adresse', 'address', 'domicile', 'residence', 'quartier', 'localisation', 'lieu', 'ville', 'commune'],
        'precedente_ecole': ['precedente_ecole', 'ecole_precedente', 'ecole_origine', 'provenance', 'ancienne_ecole', 'previous_school', 'etablissement_precedent', 'derniere_ecole', 'ecole_anterieure'],
        'date_entree': ['date_entree', 'date_d_entree', 'date_dentree', 'date entree', 'inscription', 'admission', 'date_inscription', 'entry_date', 'date_admission', 'rentree', 'date_rentree'],
        'observations': ['observations', 'obs', 'notes', 'remarques', 'commentaires', 'comment', 'remarque', 'note', 'divers'],
    }
    
    def _trouver_colonne(champ):
        """Trouve l'index d'une colonne par tous les synonymes possibles + partial match"""
        if champ not in SYNONYMES:
            return None
        for mot_cle in SYNONYMES[champ]:
            mot = _normaliser(mot_cle)
            # Exact match
            if mot in headers_bruts:
                return headers_bruts[mot]
            if mot in headers:
                return headers[mot]
            # Partial match: le mot-cle est contenu dans le header
            for k, v in headers_bruts.items():
                if mot in k or k in mot:
                    return v
            for k, v in headers.items():
                if mot in k or k in mot:
                    return v
        # Fallback ultime: cherche un mot-cle de 3+ lettres
        mots_cles_courts = [m for m in SYNONYMES[champ] if len(m) >= 3]
        for mk in mots_cles_courts:
            mk_norm = _normaliser(mk)
            for k, v in headers_bruts.items():
                if mk_norm in k or k in mk_norm:
                    return v
            for k, v in headers.items():
                if mk_norm in k or k in mk_norm:
                    return v
        return None
    
    col_prenom = _trouver_colonne('prenom')
    col_nom = _trouver_colonne('nom')
    col_sexe = _trouver_colonne('sexe')
    col_classe = _trouver_colonne('classe')
    col_tel = _trouver_colonne('tel')
    col_dn = _trouver_colonne('date_naissance')
    col_ln = _trouver_colonne('lieu_naissance')
    col_tuteur = _trouver_colonne('tuteur')
    col_adresse = _trouver_colonne('adresse')
    col_prec_ecole = _trouver_colonne('precedente_ecole')
    col_date_entree = _trouver_colonne('date_entree')
    col_obs = _trouver_colonne('observations')
    
    # Verifier les colonnes obligatoires
    if not col_prenom or not col_nom:
        cols_trouvees = ', '.join(col_names_raw[:15]) if col_names_raw else '(aucune)'
        flash(f'Colonnes PRENOM et NOM obligatoires. Colonnes detectees: {cols_trouvees}', 'danger')
        return redirect(url_for('eleves'))
    
    from datetime import datetime
    year = str(datetime.now().year)
    last_el = Eleve.query.filter_by(ecole_id=ecole_id).order_by(Eleve.id.desc()).first()
    next_id = (last_el.id + 1) if last_el else 1
    ajoutes = ignores = 0
    
    # Trouver la ligne de debut des donnees (apres les headers)
    data_start = 2
    # Verifier si la ligne 2 contient encore des headers
    if ws.cell(2, col_prenom).value:
        val_l2 = _normaliser(str(ws.cell(2, col_prenom).value))
        if val_l2 in ['prenom', 'prenoms', 'first_name', 'firstname']:
            data_start = 3
            # Verifier aussi la ligne 3
            if ws.cell(3, col_prenom).value:
                val_l3 = _normaliser(str(ws.cell(3, col_prenom).value))
                if val_l3 in ['prenom', 'prenoms', 'first_name', 'firstname']:
                    data_start = 4
    
    for row_idx in range(data_start, ws.max_row + 1):
        prenom = str(ws.cell(row_idx, col_prenom).value or '').strip()
        nom = str(ws.cell(row_idx, col_nom).value or '').strip()
        if not prenom or not nom:
            continue
        
        sexe = str(ws.cell(row_idx, col_sexe).value or 'M').strip() if col_sexe else 'M'
        sexe = sexe[0].upper() if sexe else 'M'
        
        classe_nom = str(ws.cell(row_idx, col_classe).value or '').strip() if col_classe else ''
        classe_id = None
        if classe_nom:
            cl = classes.get(classe_nom.lower())
            if not cl:
                cl = Classe(nom=classe_nom, ecole_id=ecole_id)
                db.session.add(cl); db.session.flush()
                classes[classe_nom.lower()] = cl
            classe_id = cl.id
        
        # Verification doublon : meme prenom + nom + classe (meme ecole/annee)
        existant = Eleve.query.filter_by(
            prenom=prenom, nom=nom, classe_id=classe_id, ecole_id=ecole_id, annee_scolaire=annee
        ).first()
        if existant:
            ignores += 1
            continue
        
        code = f"{year}{next_id:04d}"
        next_id += 1
        
        el = Eleve(
            code=code, prenom=prenom, nom=nom, sexe=sexe,
            classe_id=classe_id, ecole_id=ecole_id, annee_scolaire=annee,
            tel=str(ws.cell(row_idx, col_tel).value or '').strip() if col_tel else None,
            date_naissance=str(ws.cell(row_idx, col_dn).value or '').strip() if col_dn else None,
            lieu_naissance=str(ws.cell(row_idx, col_ln).value or '').strip() if col_ln else None,
            tuteur=str(ws.cell(row_idx, col_tuteur).value or '').strip() if col_tuteur else None,
            adresse=str(ws.cell(row_idx, col_adresse).value or '').strip() if col_adresse else None,
            precedente_ecole=str(ws.cell(row_idx, col_prec_ecole).value or '').strip() if col_prec_ecole else None,
            date_entree=str(ws.cell(row_idx, col_date_entree).value or '').strip() if col_date_entree else None,
            observations=str(ws.cell(row_idx, col_obs).value or '').strip() if col_obs else None,
            situation='Inscrit'
        )
        db.session.add(el)
        ajoutes += 1
    
    for cl in classes.values():
        cl.effectif = Eleve.query.filter_by(classe_id=cl.id).count()
    db.session.commit()
    
    flash(f'{ajoutes} eleve(s) importes. {ignores} ignores (doublons).', 'success')
    return redirect(url_for('eleves'))


@app.route('/eleves/modele')
@login_required
def eleves_modele():
    """Telecharger le modele Excel pour l'import d'eleves"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eleves"
    
    headers = ['Prenom', 'Nom', 'Sexe', 'Classe', 'Tel', 'Date Naissance', 'Lieu Naissance', 'Tuteur', 'Adresse', 'Precedente Ecole', 'Date Entree', 'Observations']
    
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill; cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    exemples = [
        ['Fatou', 'DIOP', 'F', 'CP1', '771234567', '15/05/2017', 'Dakar', 'Mamadou DIOP', 'Medina', 'Ecole A', '01/10/2024', ''],
        ['Abdou', 'SOW', 'M', 'CE1', '775556677', '03/12/2016', 'Thies', 'Ibrahima SOW', 'Plateau', 'Ecole B', '05/10/2024', 'Redoublant'],
    ]
    
    efill = PatternFill(start_color="E8F4E8", end_color="E8F4E8", fill_type="solid")
    for row, exemple in enumerate(exemples, 2):
        for col, val in enumerate(exemple, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if row == 2: cell.fill = efill
            cell.alignment = Alignment(horizontal='center' if col not in [1, 2, 7, 8, 10] else 'left')
    
    for i, w in enumerate([15, 15, 6, 10, 14, 16, 16, 18, 18, 20, 14, 24], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='modele_eleves.xlsx')
